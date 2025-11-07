'''
Class to handle reading from and writing to an Excel file.
'''
import openpyxl
import os
class ExcelHandler:
    def __init__(self, filename='budget.xlsx'):
        self.filename = filename
        if os.path.exists(self.filename):
            self.workbook = openpyxl.load_workbook(self.filename)
            self.expense_sheet = self.workbook["Expenses"]
            self.income_sheet = self.workbook["Income"]
        else:
            self.workbook = openpyxl.Workbook()
            self.expense_sheet = self.workbook.active
            self.expense_sheet.title = "Expenses"
            self.income_sheet = self.workbook.create_sheet(title="Income")
            # Initialize headers
            self.expense_sheet.append(["Date", "Category", "Amount"])
            self.income_sheet.append(["Date", "Source", "Amount"])
            self.workbook.save(self.filename)
    def write_expense(self, expense):
        self.expense_sheet.append([expense.date, expense.category, expense.amount])
        self.workbook.save(self.filename)  # Save after writing
    def write_income(self, income):
        self.income_sheet.append([income.date, income.source, income.amount])
        self.workbook.save(self.filename)  # Save after writing
    def save_and_close(self):
        self.workbook.save(self.filename)
        self.workbook.close()